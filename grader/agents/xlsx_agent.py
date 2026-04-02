"""Agent B: XLSX rubric evaluation (Steps 1A-1D)."""

from __future__ import annotations

import math
import re
from datetime import date, datetime
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from grader.constants import (
    ERROR_DESCRIPTIONS,
    FOR_REVIEW,
    NEEDS_MANUAL_REVIEW,
    REQUIRED_COLUMNS,
    STEP1B_REQUIRED_COLUMN_ALIASES,
    STEP1B_REQUIRED_COLUMNS,
    TRANSACTION_TYPE_ABBREVIATION_NORMALIZED,
    TRANSACTION_TYPE_ALLOWED_NORMALIZED,
    WORKSHEET_INTENTS,
)
from grader.models import ScoreValue, XlsxEvaluation
from grader.utils import (
    compact_text,
    has_time_component,
    is_blank,
    is_currency_number_format,
    is_datetime_number_format,
    is_explicit_null_marker,
    is_numeric_cell,
    is_t_id,
    looks_unsplit_merchant,
    resolve_required_columns,
    safe_str,
)


@dataclass
class TableData:
    headers: list[str]
    rows: list[dict[str, Any]]
    row_numbers: list[int]
    column_index_by_header: dict[str, int]


def run_xlsx_agent(xlsx_path: Path) -> XlsxEvaluation:
    workbook = load_workbook(xlsx_path, data_only=True)
    try:
        intent_map, intent_scores = map_sheet_intents(workbook.sheetnames)

        score_1a, notes_1a = evaluate_1a(intent_map)

        step1b_sheet = resolve_step1b_sheet(workbook)
        score_1b, errors_1b, review_flags_1b = evaluate_1b(step1b_sheet)

        report1_sheet = workbook[intent_map["report1"]] if intent_map.get("report1") else None
        report2_sheet = workbook[intent_map["report2"]] if intent_map.get("report2") else None

        score_1c, notes_1c, flags_1c = evaluate_pivot_sheet(report1_sheet, report_id=1)
        score_1d, notes_1d, flags_1d = evaluate_pivot_sheet(report2_sheet, report_id=2)

        review_flags = review_flags_1b + flags_1c + flags_1d

        # Surface low-confidence sheet intent assignments.
        for intent, score in intent_scores.items():
            if score < 78 and intent_map.get(intent):
                review_flags.append(
                    f"Worksheet intent '{intent}' matched with low confidence to sheet '{intent_map[intent]}'."
                )

        return XlsxEvaluation(
            score_1a=score_1a,
            notes_1a=notes_1a,
            score_1b=score_1b,
            errors_1b=errors_1b,
            review_flags_1b=review_flags_1b,
            score_1c=score_1c,
            notes_1c=notes_1c,
            score_1d=score_1d,
            notes_1d=notes_1d,
            review_flags=dedupe_preserve_order(review_flags),
        )
    finally:
        workbook.close()


def evaluate_1a(intent_map: dict[str, str | None]) -> tuple[float, str]:
    missing = [intent for intent, sheet in intent_map.items() if sheet is None]
    if not missing:
        return 1.0, "All seven worksheets present."
    missing_readable = ", ".join(missing)
    return 0.0, f"Missing required worksheet intents: {missing_readable}."


def evaluate_1b(step1b_sheet: Any | None) -> tuple[float, list[str], list[str]]:
    if step1b_sheet is None:
        return 0.0, ["Could not identify AllTransactionsAndCustomers-equivalent worksheet for Step 1B."], [
            "Could not confidently identify the AllTransactionsAndCustomers-equivalent worksheet for Step 1B."
        ]

    table = extract_table(step1b_sheet)
    if not table.headers:
        return 0.0, ["Step 1B target worksheet appears empty or headers are unreadable."], [
            "Unable to read Step 1B target worksheet headers with confidence."
        ]

    col, missing, _, ambiguous_required = resolve_required_columns(
        table.headers,
        STEP1B_REQUIRED_COLUMNS,
        STEP1B_REQUIRED_COLUMN_ALIASES,
    )

    errors: list[str] = []
    review_flags: list[str] = []

    # COLUMNS
    if missing:
        parts: list[str] = [ERROR_DESCRIPTIONS["COLUMNS"]]
        parts.append(f"Missing: {', '.join(missing)}")
        errors.append(" ".join(parts))

    if ambiguous_required:
        review_flags.append(
            "Column matching was ambiguous for: "
            + ", ".join(ambiguous_required)
            + ". Resolved using closest normalized header match."
        )

    # DUPLICATES
    dup_issue, dup_note = check_duplicates(table.rows, col)
    if dup_issue:
        errors.append(dup_note)

    # TRANSACTION TYPE VALUES
    ttype_issue, ttype_note, ttype_review = check_transaction_type(table.rows, col)
    if ttype_issue:
        errors.append(ttype_note)
    review_flags.extend(ttype_review)

    # TRANSACTION ID FORMAT
    id_issue, id_note, id_review = check_transaction_id(table.rows, col)
    if id_issue:
        errors.append(id_note)
    review_flags.extend(id_review)

    # NAME FORMAT
    name_issue, name_note, name_review = check_name_format(table.rows, col)
    if name_issue:
        errors.append(name_note)
    review_flags.extend(name_review)

    # OPEN DATE TIME FORMAT
    open_date_issue, open_date_note, open_date_review = check_open_date_time(step1b_sheet, table)
    if open_date_issue:
        errors.append(open_date_note)
    review_flags.extend(open_date_review)

    # MERCHANT COLUMN
    merchant_issue, merchant_note, merchant_review = check_merchant(table.rows, col)
    if merchant_issue:
        errors.append(merchant_note)
    review_flags.extend(merchant_review)

    # CITY/STATE COLUMNS
    city_state_issue, city_state_note, city_state_review = check_city_state(table.rows, col)
    if city_state_issue:
        errors.append(city_state_note)
    review_flags.extend(city_state_review)

    # TOTAL VALUE
    total_issue, total_note, total_review = check_total_value(step1b_sheet, table, col)
    if total_issue:
        errors.append(total_note)
    review_flags.extend(total_review)

    # NOTES COLUMN
    notes_issue, notes_note, notes_review = check_notes(table.rows, col)
    if notes_issue:
        errors.append(notes_note)
    review_flags.extend(notes_review)

    # NULL VALUES
    null_issue, null_note, null_review = check_null_values(table.rows, col)
    if null_issue:
        errors.append(null_note)
    review_flags.extend(null_review)

    score = max(0.0, 4.0 - 0.5 * len(errors))

    if not errors:
        errors = ["No errors found."]

    return score, errors, dedupe_preserve_order(review_flags)


def evaluate_pivot_sheet(ws: Any | None, report_id: int) -> tuple[ScoreValue, str, list[str]]:
    if ws is None:
        return 0.0, f"Report {report_id} worksheet is missing.", []

    pivots = getattr(ws, "_pivots", None)
    if pivots is None:
        return FOR_REVIEW, (
            f"{NEEDS_MANUAL_REVIEW}: ws._pivots unavailable in this environment for Report {report_id}."
        ), [f"Report {report_id}: pivot object verification unavailable."]

    if len(pivots) == 0:
        return 0.0, "No real PivotTable object found (ws._pivots is empty).", []

    field_check = check_pivot_fields(pivots[0], report_id)
    values_ok = pivot_has_values(ws, pivots[0])

    if field_check is None:
        return FOR_REVIEW, (
            f"{NEEDS_MANUAL_REVIEW}: pivot object exists, but field configuration could not be verified confidently."
        ), [f"Report {report_id}: unable to inspect pivot field configuration with confidence."]

    if field_check and values_ok:
        return 0.5, "Pivot object exists and required dimensions/value field checks passed.", []

    reasons = []
    if not field_check:
        reasons.append("pivot fields/aggregation do not match required configuration")
    if not values_ok:
        reasons.append("pivot values appear empty")
    return 0.0, "Pivot validation failed: " + "; ".join(reasons) + ".", []


def map_sheet_intents(sheet_names: list[str]) -> tuple[dict[str, str | None], dict[str, int]]:
    scores: dict[str, list[tuple[int, str]]] = {intent: [] for intent in WORKSHEET_INTENTS}
    for sheet_name in sheet_names:
        for intent in WORKSHEET_INTENTS:
            score = score_sheet_for_intent(sheet_name, intent)
            scores[intent].append((score, sheet_name))

    intent_map: dict[str, str | None] = {}
    intent_scores: dict[str, int] = {}
    for intent, options in scores.items():
        options = sorted(options, key=lambda x: (x[0], x[1]))
        best_score, best_sheet = options[-1]
        if best_score >= 70:
            intent_map[intent] = best_sheet
            intent_scores[intent] = best_score
        else:
            intent_map[intent] = None
            intent_scores[intent] = best_score

    # Resolve frequent overlap between merged and cleaned transaction sheets.
    merged = intent_map.get("all_transactions_customers")
    cleaned = intent_map.get("all_transactions_cleaned")
    if merged and cleaned and merged == cleaned:
        alternatives = sorted(scores["all_transactions_cleaned"], key=lambda x: (x[0], x[1]), reverse=True)
        replacement = None
        replacement_score = 0
        for score, candidate in alternatives:
            if candidate != merged and score >= 65:
                replacement = candidate
                replacement_score = score
                break
        if replacement:
            intent_map["all_transactions_cleaned"] = replacement
            intent_scores["all_transactions_cleaned"] = replacement_score

    return intent_map, intent_scores


def score_sheet_for_intent(sheet_name: str, intent: str) -> int:
    compact_name = compact_text(sheet_name)
    aliases = [compact_text(a) for a in WORKSHEET_INTENTS[intent]]

    best = 0
    for alias in aliases:
        if not alias:
            continue
        if alias == compact_name:
            return 100
        if alias in compact_name or compact_name in alias:
            best = max(best, 95)
        ratio = int(SequenceMatcher(None, compact_name, alias).ratio() * 100)
        best = max(best, ratio)

    # Intent-specific nudges to avoid false positives.
    if intent == "all_transactions_cleaned" and "customer" in compact_name:
        best -= 25
    if intent == "all_transactions_customers" and "customer" not in compact_name:
        best -= 20
    if intent == "report1" and "2" in compact_name:
        best -= 25
    if intent == "report2" and "1" in compact_name:
        best -= 25

    return max(0, min(100, best))


def resolve_step1b_sheet(workbook: Any) -> Any | None:
    required_identifier = "alltransactionsandcustomers"
    legacy_identifier = "alltransactionsandcustomer"
    preferred_name = "alltransactionsandcustomersdata"
    legacy_preferred_name = "alltransactionsandcustomerdata"

    matches: list[tuple[int, int, int, Any]] = []
    for index, ws in enumerate(workbook.worksheets):
        compact_name = compact_text(ws.title)
        if required_identifier not in compact_name and legacy_identifier not in compact_name:
            continue

        score = 0
        if compact_name == preferred_name or compact_name == legacy_preferred_name:
            score += 4
        if preferred_name in compact_name or legacy_preferred_name in compact_name:
            score += 2
        if compact_name.startswith(required_identifier) or compact_name.startswith(legacy_identifier):
            score += 1

        # Prefer deterministic first-seen worksheet order when scores tie.
        matches.append((score, -len(compact_name), -index, ws))

    if not matches:
        return None

    matches.sort(reverse=True)
    return matches[0][3]


def resolve_cleaned_sheet(workbook: Any, intent_map: dict[str, str | None]) -> Any | None:
    cleaned_name = intent_map.get("all_transactions_cleaned")
    if cleaned_name:
        return workbook[cleaned_name]

    required_canon = {compact_text(c) for c in REQUIRED_COLUMNS}
    best_sheet = None
    best_score = -1
    for ws in workbook.worksheets:
        table = extract_table(ws)
        header_set = {compact_text(h) for h in table.headers}
        overlap = len(required_canon.intersection(header_set))
        if overlap > best_score and "customer" not in compact_text(ws.title):
            best_score = overlap
            best_sheet = ws

    return best_sheet if best_score >= 7 else None


def extract_table(ws: Any) -> TableData:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return TableData(headers=[], rows=[], row_numbers=[], column_index_by_header={})

    header_row = detect_header_row(ws)
    raw_headers = [ws.cell(header_row, c).value for c in range(1, max_col + 1)]
    active_indices = [
        i for i, value in enumerate(raw_headers, start=1)
        if not is_blank(value)
    ]

    headers = [safe_str(raw_headers[i - 1]).strip() for i in active_indices]
    header_index = {headers[pos]: active_indices[pos] for pos in range(len(headers))}

    rows: list[dict[str, Any]] = []
    row_numbers: list[int] = []
    for row_idx in range(header_row + 1, max_row + 1):
        values = [ws.cell(row_idx, c).value for c in active_indices]
        if all(is_blank(v) for v in values):
            continue
        row = {headers[pos]: values[pos] for pos in range(len(headers))}
        rows.append(row)
        row_numbers.append(row_idx)

    return TableData(
        headers=headers,
        rows=rows,
        row_numbers=row_numbers,
        column_index_by_header=header_index,
    )


def detect_header_row(ws: Any) -> int:
    max_col = ws.max_column or 0
    limit = min(ws.max_row or 1, 15)
    best_row = 1
    best_score = -1

    column_hints = {compact_text(c) for c in REQUIRED_COLUMNS}
    for row_idx in range(1, limit + 1):
        values = [ws.cell(row_idx, c).value for c in range(1, max_col + 1)]
        non_blank = [v for v in values if not is_blank(v)]
        hint_hits = sum(1 for v in non_blank if compact_text(safe_str(v)) in column_hints)
        score = len(non_blank) + 2 * hint_hits
        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def check_duplicates(rows: list[dict[str, Any]], col_map: dict[str, str | None]) -> tuple[bool, str]:
    needed = [col_map[c] for c in REQUIRED_COLUMNS if col_map.get(c)]
    if len(needed) < 5:
        return False, ""

    tx_col = col_map.get("TransactionID")
    seen = set()
    duplicate_count = 0
    duplicate_ids: set[str] = set()
    for row in rows:
        sig_values = [normalize_cell(row.get(c)) for c in needed]
        # Ignore fully blank rows while preserving exact-row duplicate detection.
        if all(value == "" for value in sig_values):
            continue
        sig = tuple(sig_values)
        if sig in seen:
            duplicate_count += 1
            if tx_col:
                tid = safe_str(row.get(tx_col)).strip().upper()
                if tid:
                    duplicate_ids.add(tid)
        else:
            seen.add(sig)

    if duplicate_count > 0:
        note = f"{ERROR_DESCRIPTIONS['DUPLICATES']} Found {duplicate_count} duplicate row(s)."
        if duplicate_ids:
            preview = ", ".join(sorted(duplicate_ids)[:5])
            note += f" Sample duplicate TransactionID values: {preview}."
        return True, note

    return False, ""


def check_transaction_type(rows: list[dict[str, Any]], col_map: dict[str, str | None]) -> tuple[bool, str, list[str]]:
    col = col_map.get("TransactionType")
    if not col:
        return False, "", ["Cannot verify TransactionType values because column is missing."]

    invalid: set[str] = set()
    abbreviations_found: set[str] = set()
    other_invalid_found: set[str] = set()
    invalid_counts: dict[str, int] = {}
    for row in rows:
        value = safe_str(row.get(col)).strip()
        if not value:
            continue
        normalized = normalize_tx_type(value)
        if normalized in TRANSACTION_TYPE_ALLOWED_NORMALIZED:
            continue
        if normalized in TRANSACTION_TYPE_ABBREVIATION_NORMALIZED:
            abbreviations_found.add(value)
            invalid.add(value)
            invalid_counts[value] = invalid_counts.get(value, 0) + 1
            continue
        other_invalid_found.add(value)
        invalid.add(value)
        invalid_counts[value] = invalid_counts.get(value, 0) + 1

    if invalid:
        total_invalid_rows = sum(invalid_counts.values())

        def format_with_counts(values: set[str]) -> str:
            ranked = sorted(values, key=lambda v: (-invalid_counts.get(v, 0), v))[:5]
            return ", ".join(f"{v} ({invalid_counts.get(v, 0)})" for v in ranked)

        details: list[str] = []
        if abbreviations_found:
            details.append(f"abbreviations: {format_with_counts(abbreviations_found)}")
        if other_invalid_found:
            details.append(f"other values: {format_with_counts(other_invalid_found)}")

        message = (
            f"{ERROR_DESCRIPTIONS['TRANSACTION_TYPE_VALUES']} "
            "Accepted values are full words: Fee, Deposit, Withdrawal, Transfer, Payment, Transaction. "
            f"Detected {total_invalid_rows} invalid TransactionType row(s)."
        )
        if abbreviations_found:
            message += " Full-word values are accepted; only the listed invalid rows triggered this deduction."
        if details:
            message += " Found " + "; ".join(details) + "."

        return True, (
            message
        ), []
    return False, "", []


def check_transaction_id(rows: list[dict[str, Any]], col_map: dict[str, str | None]) -> tuple[bool, str, list[str]]:
    col = col_map.get("TransactionID")
    if not col:
        return False, "", ["Cannot verify TransactionID format because column is missing."]

    invalid = []
    for row in rows:
        value = row.get(col)
        if is_blank(value):
            continue
        if not is_t_id(value):
            invalid.append(safe_str(value).strip())

    if invalid:
        preview = ", ".join(invalid[:5])
        return True, f"{ERROR_DESCRIPTIONS['TRANSACTION_ID_FORMAT']} Sample invalid IDs: {preview}.", []
    return False, "", []


def check_name_format(rows: list[dict[str, Any]], col_map: dict[str, str | None]) -> tuple[bool, str, list[str]]:
    first_col = col_map.get("CustomerFName")
    last_col = col_map.get("CustomerLName")
    if not first_col or not last_col:
        return False, "", ["Cannot fully verify name format because first/last name columns are missing."]

    issues = 0
    for row in rows:
        if has_unsplit_name_issue(row.get(first_col), row.get(last_col)):
            issues += 1

    if issues > 0:
        return True, f"{ERROR_DESCRIPTIONS['NAME_FORMAT']} Found {issues} row(s) with unsplit name patterns.", []
    return False, "", []


def resolve_open_date_column(headers: list[str]) -> str | None:
    mapping, _, _, _ = resolve_required_columns(headers, ["OpenDate"])
    direct = mapping.get("OpenDate")
    if direct:
        return direct

    fallback_candidates: list[tuple[int, str]] = []
    for header in headers:
        compact_header = compact_text(header)
        if compact_header == "date":
            fallback_candidates.append((1, header))
            continue
        if "opendate" in compact_header:
            fallback_candidates.append((0, header))

    if not fallback_candidates:
        return None

    fallback_candidates.sort(key=lambda item: (item[0], item[1].lower()))
    return fallback_candidates[0][1]


def check_open_date_time(ws: Any, table: TableData) -> tuple[bool, str, list[str]]:
    open_date_col = resolve_open_date_column(table.headers)
    if not open_date_col:
        return False, "", []

    col_idx = table.column_index_by_header.get(open_date_col)
    missing_time_issues = 0
    for idx, row in enumerate(table.rows):
        value = row.get(open_date_col)
        if is_blank(value):
            continue

        if has_time_component(value) or has_time_in_excel_serial(value):
            continue

        number_format = ""
        if col_idx is not None and idx < len(table.row_numbers):
            row_idx = table.row_numbers[idx]
            number_format = safe_str(ws.cell(row_idx, col_idx).number_format)

        if is_datetime_number_format(number_format):
            continue

        missing_time_issues += 1

    if missing_time_issues > 0:
        return (
            True,
            f"{ERROR_DESCRIPTIONS['OPEN_DATE_TIME']} Found {missing_time_issues} row(s) without a time component.",
            [],
        )

    return False, "", []


def check_merchant(rows: list[dict[str, Any]], col_map: dict[str, str | None]) -> tuple[bool, str, list[str]]:
    col = col_map.get("Merchant")
    if not col:
        return False, "", ["Cannot verify Merchant cleanup because Merchant column is missing."]

    issues = 0
    for row in rows:
        merchant = row.get(col)
        if is_blank(merchant):
            continue
        if looks_unsplit_merchant(merchant):
            issues += 1

    if issues > 0:
        return True, f"{ERROR_DESCRIPTIONS['MERCHANT_COLUMN']} Found {issues} unsplit merchant row(s).", []
    return False, "", []


def check_city_state(rows: list[dict[str, Any]], col_map: dict[str, str | None]) -> tuple[bool, str, list[str]]:
    merchant_col = col_map.get("Merchant")
    city_col = col_map.get("City")
    state_col = col_map.get("State")

    missing_cols = [name for name, c in [("Merchant", merchant_col), ("City", city_col), ("State", state_col)] if c is None]
    if missing_cols:
        return False, "", [f"Cannot verify City/State split because column(s) missing: {', '.join(missing_cols)}."]

    issues = 0
    for row in rows:
        merchant = row.get(merchant_col)
        if looks_unsplit_merchant(merchant):
            # Avoid double counting when merchant carries city/state text.
            continue

        if is_blank(merchant):
            continue

        city_text = safe_str(row.get(city_col)).strip()
        state_text = safe_str(row.get(state_col)).strip()

        # Blank city/state are allowed. Only flag rows where a combined
        # "City, ST" value appears unsplit in City while State is empty.
        if city_text and not state_text and re.search(r",\s*[A-Za-z]{2}\s*$", city_text):
            issues += 1

    if issues > 0:
        return True, f"{ERROR_DESCRIPTIONS['CITY_STATE_COLUMNS']} Found {issues} row(s) with city/state split issues.", []
    return False, "", []


def check_total_value(ws: Any, table: TableData, col_map: dict[str, str | None]) -> tuple[bool, str, list[str]]:
    col = col_map.get("TotalValue")
    if not col:
        return False, "", ["Cannot verify TotalValue typing because TotalValue column is missing."]

    col_idx = table.column_index_by_header.get(col)
    numeric_issues = 0
    currency_issues = 0
    explicit_format_rows = 0
    general_format_rows = 0
    ambiguous_general_rows = 0

    for idx, row in enumerate(table.rows):
        value = row.get(col)
        if is_blank(value):
            continue
        if not is_numeric_cell(value):
            numeric_issues += 1
            continue

        number_format = ""
        if col_idx is not None and idx < len(table.row_numbers):
            row_idx = table.row_numbers[idx]
            number_format = safe_str(ws.cell(row_idx, col_idx).number_format)

        if number_format and number_format.lower() != "general":
            explicit_format_rows += 1
            if not is_currency_number_format(number_format):
                currency_issues += 1
        else:
            general_format_rows += 1
            if not looks_currency_like_numeric(value):
                ambiguous_general_rows += 1

    if numeric_issues > 0 or currency_issues > 0:
        parts: list[str] = []
        if numeric_issues > 0:
            parts.append(f"{numeric_issues} row(s) are non-numeric")
        if currency_issues > 0:
            parts.append(f"{currency_issues} row(s) are not currency-formatted")
        return (
            True,
            f"{ERROR_DESCRIPTIONS['TOTAL_VALUE']} Detected "
            + " and ".join(parts)
            + ". This is applied as one criterion-level deduction.",
            [],
        )

    review_flags: list[str] = []
    if ambiguous_general_rows > 0:
        review_flags.append(
            f"Currency formatting could not be confidently verified for {ambiguous_general_rows} TotalValue row(s)."
        )

    return False, "", review_flags


def check_notes(rows: list[dict[str, Any]], col_map: dict[str, str | None]) -> tuple[bool, str, list[str]]:
    col = col_map.get("Notes")
    if not col:
        return False, "", ["Cannot verify Notes cleanup because Notes column is missing."]

    issues = 0
    error_fragments = ["#error", "#n/a", "#value", "#div", "#ref", "#name"]
    for row in rows:
        text = safe_str(row.get(col))
        lower = text.lower()
        has_qerr = bool(re.search(r"\?+\s*err(or)?", lower))
        has_sheet_error = any(fragment in lower for fragment in error_fragments)
        if "@" in text or has_qerr or has_sheet_error:
            issues += 1

    if issues > 0:
        return True, f"{ERROR_DESCRIPTIONS['NOTES_COLUMN']} Found {issues} row(s) with unresolved note artifacts.", []
    return False, "", []


def check_null_values(rows: list[dict[str, Any]], col_map: dict[str, str | None]) -> tuple[bool, str, list[str]]:
    missing_required_cols = [c for c in REQUIRED_COLUMNS if col_map.get(c) is None]
    if missing_required_cols:
        return False, "", [
            "Cannot fully verify null-value criterion because required columns are missing."
        ]

    issues = 0
    for row in rows:
        tx_type = normalize_tx_type(row.get(col_map["TransactionType"]))
        merchant_val = row.get(col_map["Merchant"])
        merchant_unsplit = looks_unsplit_merchant(merchant_val)

        for required_col in REQUIRED_COLUMNS:
            if required_col == "Notes":
                # Notes can be blank and are checked separately for artifacts.
                continue
            value = row.get(col_map[required_col])
            if required_col in {"Merchant", "City", "State"}:
                if tx_type in {"fee", "transfer", "tra"}:
                    continue
                if merchant_unsplit and required_col in {"City", "State"}:
                    # avoid double counting with Merchant criterion
                    continue
            if is_explicit_null_marker(value):
                issues += 1
                break

    if issues > 0:
        return True, f"{ERROR_DESCRIPTIONS['NULL_VALUES']} Found {issues} row(s) with explicit null markers.", []
    return False, "", []


def check_pivot_fields(pivot: Any, report_id: int) -> bool | None:
    names = extract_pivot_field_names(pivot)
    data_fields = extract_pivot_data_fields(pivot)

    if names is None and data_fields is None:
        return None

    names = names or set()
    data_fields = data_fields or []

    if report_id == 1:
        has_customer_axis = any("customerlname" in n or "customerfname" in n or "customer" in n for n in names)
        has_tx_axis = any("transactiontype" in n for n in names)
        has_value = any("totalvalue" in n and (agg in {"sum", "summarize"} or agg == "") for n, agg in data_fields)
        return has_customer_axis and has_tx_axis and has_value

    has_tx_axis = any("transactiontype" in n for n in names)
    has_status_axis = any("accountstatuslevel" in n or "status" in n for n in names)
    has_value = any("totalvalue" in n and (agg in {"average", "avg"} or agg == "") for n, agg in data_fields)
    return has_tx_axis and has_status_axis and has_value


def extract_pivot_field_names(pivot: Any) -> set[str] | None:
    names: set[str] = set()
    try:
        cache = getattr(pivot, "cache", None)
        cache_fields = getattr(cache, "cacheFields", None)
        if cache_fields is not None:
            for field in cache_fields:
                name = getattr(field, "name", None)
                if name:
                    names.add(compact_text(name))

        for collection_name in ["pivotFields", "rowFields", "colFields"]:
            collection = getattr(pivot, collection_name, None)
            if collection is None:
                continue
            for item in iterable_of(collection):
                for attr in ["name", "x", "field"]:
                    value = getattr(item, attr, None)
                    if isinstance(value, str) and value:
                        names.add(compact_text(value))
        return names
    except Exception:  # noqa: BLE001
        return None


def extract_pivot_data_fields(pivot: Any) -> list[tuple[str, str]] | None:
    results: list[tuple[str, str]] = []
    try:
        data_fields = getattr(pivot, "dataFields", None)
        if data_fields is None:
            return []
        for item in iterable_of(data_fields):
            name = compact_text(safe_str(getattr(item, "name", "")))
            agg = compact_text(safe_str(getattr(item, "subtotal", "")))
            results.append((name, agg))
        return results
    except Exception:  # noqa: BLE001
        return None


def pivot_has_values(ws: Any, pivot: Any) -> bool:
    try:
        location = getattr(pivot, "location", None)
        ref = getattr(location, "ref", None)
        if ref:
            numeric = 0
            for row in ws[ref]:
                for cell in row:
                    if is_numeric_cell(cell.value):
                        numeric += 1
                        if numeric >= 3:
                            return True
            return False
    except Exception:  # noqa: BLE001
        pass

    # Fallback: inspect used range for numeric content.
    numeric = 0
    max_row = min(ws.max_row or 0, 200)
    max_col = min(ws.max_column or 0, 30)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            value = ws.cell(r, c).value
            if is_numeric_cell(value):
                numeric += 1
                if numeric >= 5:
                    return True
    return False


def iterable_of(collection: Any) -> list[Any]:
    try:
        return list(collection)
    except Exception:  # noqa: BLE001
        return []


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.10f}".rstrip("0").rstrip(".")
    return safe_str(value).strip().lower()


def normalize_tx_type(value: Any) -> str:
    text = safe_str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    # Treat trailing punctuation/noise as formatting artifacts, not semantic value changes.
    text = re.sub(r"^[^a-z]+", "", text)
    text = re.sub(r"[^a-z]+$", "", text)
    return text


def is_date_like_without_time(value: Any) -> bool:
    if isinstance(value, datetime):
        return value.time() == datetime.min.time()
    if isinstance(value, date):
        return True
    if looks_like_excel_date_serial(value):
        return not has_time_in_excel_serial(value)
    if isinstance(value, str):
        return bool(re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", value.strip()))
    return False


def looks_currency_like_numeric(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return value == 0
    if isinstance(value, float):
        if not math.isfinite(value):
            return False
        if abs(value) < 1e-9:
            return True

        # For General-format values, require cent precision and avoid
        # one-decimal/integer-like numbers that are often ambiguous.
        cents_aligned = abs((value * 100) - round(value * 100)) < 1e-6
        tenths_aligned = abs((value * 10) - round(value * 10)) < 1e-6
        return cents_aligned and not tenths_aligned
    return False


def looks_like_excel_date_serial(value: Any) -> bool:
    if not is_numeric_cell(value):
        return False
    serial = float(value)
    if not math.isfinite(serial):
        return False
    return 20000 <= serial <= 70000


def has_time_in_excel_serial(value: Any) -> bool:
    if not looks_like_excel_date_serial(value):
        return False
    serial = float(value)
    whole_days = round(serial)
    return abs(serial - whole_days) > 1e-9


def has_unsplit_name_issue(first_name: Any, last_name: Any) -> bool:
    first = safe_str(first_name).strip()
    last = safe_str(last_name).strip()

    if first == "" and last == "":
        return False

    # Clear unsplit signals: delimiters or duplicated full-name strings.
    if "," in first and last == "":
        return True
    if " - " in first and last == "":
        return True

    first_tokens = alpha_tokens(first)
    last_tokens = alpha_tokens(last)

    if len(first_tokens) >= 2 and last == "":
        return True

    if first != "" and last != "" and first.lower() == last.lower() and len(first_tokens) >= 2:
        return True

    if first == "" and len(last_tokens) >= 2:
        return True

    return False


def alpha_tokens(text: str) -> list[str]:
    return [token for token in re.split(r"[\s\-']+", text) if re.search(r"[A-Za-z]", token)]


def dedupe_preserve_order(values: list[str]) -> list[str]:
    seen = set()
    result = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result
